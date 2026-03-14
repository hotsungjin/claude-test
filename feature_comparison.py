import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "기능 비교표"

# ── 색상 정의 ──
COLOR_HEADER_BG   = "1F4E79"  # 진파랑
COLOR_CATEGORY_BG = "2E75B6"  # 중간파랑
COLOR_YES         = "C6EFCE"  # 연초록
COLOR_PARTIAL     = "FFEB9C"  # 연노랑
COLOR_NO          = "FFC7CE"  # 연빨강
COLOR_NA          = "EDEDED"  # 연회색

FILL_HEADER   = PatternFill("solid", fgColor=COLOR_HEADER_BG)
FILL_CATEGORY = PatternFill("solid", fgColor=COLOR_CATEGORY_BG)
FILL_YES      = PatternFill("solid", fgColor=COLOR_YES)
FILL_PARTIAL  = PatternFill("solid", fgColor=COLOR_PARTIAL)
FILL_NO       = PatternFill("solid", fgColor=COLOR_NO)
FILL_NA       = PatternFill("solid", fgColor=COLOR_NA)

FONT_HEADER   = Font(bold=True, color="FFFFFF", size=11)
FONT_CATEGORY = Font(bold=True, color="FFFFFF", size=10)
FONT_NORMAL   = Font(size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 헤더 행 ──
headers = ["카테고리", "고도몰 기능명", "포함여부", "비고", "설성목장 필요여부", "우선순위", "담당자 메모"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = FILL_HEADER
    cell.font = FONT_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER

# ── 데이터 ──
# (카테고리, 기능명, 포함여부, 비고)
data = [
    # 1. 회원 관리
    ("1. 회원 관리", "회원 가입/탈퇴/수정", "✅ 포함", ""),
    ("1. 회원 관리", "회원 목록/검색/필터", "✅ 포함", ""),
    ("1. 회원 관리", "회원 그룹/등급제", "✅ 포함", ""),
    ("1. 회원 관리", "회원 휴면 처리", "❌ 미포함", "개인정보보호법 의무사항"),
    ("1. 회원 관리", "회원 마일리지 관리", "✅ 포함", ""),
    ("1. 회원 관리", "마일리지 만료 처리", "❌ 미포함", ""),
    ("1. 회원 관리", "회원 예치금(캐시)", "❌ 미포함", ""),
    ("1. 회원 관리", "회원 CRM 통합뷰", "🔶 부분포함", "기본만"),
    ("1. 회원 관리", "본인인증 (휴대폰/I-PIN)", "❌ 미포함", "법적 의무 가능성"),
    ("1. 회원 관리", "사업자 인증", "❌ 미포함", ""),
    ("1. 회원 관리", "소셜 로그인 (카카오/네이버/구글/애플)", "✅ 포함", ""),
    ("1. 회원 관리", "PAYCO 로그인", "➖ 불필요", ""),
    ("1. 회원 관리", "회원 엑셀 일괄 등록/내보내기", "❌ 미포함", ""),
    ("1. 회원 관리", "회원 가입 이벤트 설정", "❌ 미포함", ""),
    ("1. 회원 관리", "앱 디바이스 추적", "❌ 미포함", ""),

    # 2. 상품 관리
    ("2. 상품 관리", "상품 등록/수정/삭제", "✅ 포함", ""),
    ("2. 상품 관리", "상품 카테고리 트리", "✅ 포함", ""),
    ("2. 상품 관리", "상품 옵션 (단독/조합)", "🔶 부분포함", "단순 옵션만"),
    ("2. 상품 관리", "상품 재고 관리", "✅ 포함", ""),
    ("2. 상품 관리", "품절 알림 (관리자/회원)", "✅ 포함", ""),
    ("2. 상품 관리", "상품 아이콘/뱃지", "❌ 미포함", ""),
    ("2. 상품 관리", "상품 이미지 일괄 업로드", "🔶 부분포함", ""),
    ("2. 상품 관리", "상품 일괄 가격/재고 수정", "❌ 미포함", ""),
    ("2. 상품 관리", "추가 상품 (add-on)", "❌ 미포함", ""),
    ("2. 상품 관리", "묶음 상품", "❌ 미포함", ""),
    ("2. 상품 관리", "사은품 설정", "❌ 미포함", ""),
    ("2. 상품 관리", "공통 상품 콘텐츠 템플릿", "❌ 미포함", ""),
    ("2. 상품 관리", "상품 필수 정보 (법정표시)", "❌ 미포함", "법적 의무"),
    ("2. 상품 관리", "연관 상품 / 추천 상품", "❌ 미포함", ""),
    ("2. 상품 관리", "상품 엑셀 일괄 등록", "❌ 미포함", ""),
    ("2. 상품 관리", "네이버 브랜드 인증", "❌ 미포함", ""),
    ("2. 상품 관리", "상품 수정 이력 로그", "❌ 미포함", ""),
    ("2. 상품 관리", "공급사(SCM) 상품 연동", "❌ 미포함", ""),

    # 3. 주문 관리
    ("3. 주문 관리", "주문 목록/상세 조회", "✅ 포함", ""),
    ("3. 주문 관리", "주문 상태별 관리", "✅ 포함", ""),
    ("3. 주문 관리", "주문 직접 등록 (수기주문)", "❌ 미포함", ""),
    ("3. 주문 관리", "주문 취소/반품/교환/환불", "🔶 부분포함", "기본 취소만"),
    ("3. 주문 관리", "부분 취소/부분 환불", "❌ 미포함", ""),
    ("3. 주문 관리", "다중 배송지 주문", "❌ 미포함", ""),
    ("3. 주문 관리", "주문 인쇄 (납품서/거래명세서)", "❌ 미포함", ""),
    ("3. 주문 관리", "현금영수증 발행", "❌ 미포함", "법적 의무"),
    ("3. 주문 관리", "세금계산서 발행", "❌ 미포함", "Phase 3 예정"),
    ("3. 주문 관리", "에스크로 관리", "❌ 미포함", "특정 PG 의무"),
    ("3. 주문 관리", "가상계좌/무통장입금", "❌ 미포함", ""),
    ("3. 주문 관리", "은행 데이터 자동 매칭", "❌ 미포함", ""),
    ("3. 주문 관리", "주문 메모/관리자 메모", "❌ 미포함", ""),
    ("3. 주문 관리", "주문 추가 필드 (커스텀)", "❌ 미포함", ""),
    ("3. 주문 관리", "주문 엑셀 내보내기", "❌ 미포함", ""),
    ("3. 주문 관리", "쿠폰 주문 적용/복원", "🔶 부분포함", ""),
    ("3. 주문 관리", "네이버페이 주문 관리", "❌ 미포함", ""),

    # 4. 결제
    ("4. 결제", "토스페이먼츠", "✅ 포함", ""),
    ("4. 결제", "카카오페이", "🔶 부분포함", "토스 위젯 통해 가능"),
    ("4. 결제", "네이버페이", "❌ 미포함", "직접 연동 별도 필요"),
    ("4. 결제", "신용카드 (이니시스/KCP/나이스)", "🔶 부분포함", "토스 통합"),
    ("4. 결제", "무통장입금", "❌ 미포함", ""),
    ("4. 결제", "실시간 계좌이체", "🔶 부분포함", "토스 통해 가능"),
    ("4. 결제", "해외결제 PG", "❌ 미포함", ""),
    ("4. 결제", "무이자 할부 설정", "❌ 미포함", ""),
    ("4. 결제", "결제 부분취소", "❌ 미포함", ""),
    ("4. 결제", "결제 이력 조회", "🔶 부분포함", ""),
    ("4. 결제", "PG사 20개+ 선택", "❌ 미포함", "토스 단일로 통합"),

    # 5. 배송/물류
    ("5. 배송/물류", "택배사 관리", "🔶 부분포함", "기본"),
    ("5. 배송/물류", "배송비 정책 (무료/유료/조건부)", "🔶 부분포함", ""),
    ("5. 배송/물류", "지역별 추가 배송비 (제주/도서산간)", "❌ 미포함", ""),
    ("5. 배송/물류", "배송 추적", "✅ 포함", "Phase 2"),
    ("5. 배송/물류", "다중 배송지", "❌ 미포함", ""),
    ("5. 배송/물류", "해외 배송", "❌ 미포함", ""),
    ("5. 배송/물류", "묶음 배송 설정", "❌ 미포함", ""),
    ("5. 배송/물류", "고도우체 연동", "➖ 불필요", "자체 택배사 API로 대체"),
    ("5. 배송/물류", "반품/교환 배송 관리", "❌ 미포함", ""),
    ("5. 배송/물류", "입고 관리", "❌ 미포함", ""),

    # 6. 마케팅/프로모션
    ("6. 마케팅/프로모션", "쿠폰 발행/관리", "✅ 포함", "Phase 2"),
    ("6. 마케팅/프로모션", "쿠폰 종류 (정액/정률/배송비)", "🔶 부분포함", ""),
    ("6. 마케팅/프로모션", "오프라인 쿠폰", "❌ 미포함", ""),
    ("6. 마케팅/프로모션", "타임세일", "✅ 포함", "Phase 3"),
    ("6. 마케팅/프로모션", "이벤트 특가", "🔶 부분포함", ""),
    ("6. 마케팅/프로모션", "출석 이벤트", "❌ 미포함", ""),
    ("6. 마케팅/프로모션", "장바구니 리마인더", "❌ 미포함", ""),
    ("6. 마케팅/프로모션", "재방문 쿠폰 (컴백쿠폰)", "❌ 미포함", ""),
    ("6. 마케팅/프로모션", "QR코드 생성", "❌ 미포함", ""),
    ("6. 마케팅/프로모션", "단축 URL", "❌ 미포함", ""),
    ("6. 마케팅/프로모션", "바코드 쿠폰", "❌ 미포함", ""),
    ("6. 마케팅/프로모션", "설문/투표 시스템", "❌ 미포함", ""),
    ("6. 마케팅/프로모션", "추천 리워드 시스템", "✅ 포함", "기존 sulsung-reward 통합"),
    ("6. 마케팅/프로모션", "SNS 공유 설정", "❌ 미포함", ""),
    ("6. 마케팅/프로모션", "보험 위젯", "➖ 불필요", ""),

    # 7. 통계/분석
    ("7. 통계/분석", "방문자 통계 (일/주/월/시간)", "❌ 미포함", "Vercel Analytics로 대체"),
    ("7. 통계/분석", "유입 경로 분석", "❌ 미포함", ""),
    ("7. 통계/분석", "검색어 분석", "❌ 미포함", ""),
    ("7. 통계/분석", "회원 통계 (연령/성별/지역)", "❌ 미포함", ""),
    ("7. 통계/분석", "상품별 판매 통계", "✅ 포함", "Phase 2"),
    ("7. 통계/분석", "주문 통계", "✅ 포함", ""),
    ("7. 통계/분석", "매출 통계 (일/주/월)", "✅ 포함", ""),
    ("7. 통계/분석", "장바구니 분석", "❌ 미포함", ""),
    ("7. 통계/분석", "위시리스트 분석", "❌ 미포함", ""),
    ("7. 통계/분석", "공급사별 정산 통계", "❌ 미포함", ""),
    ("7. 통계/분석", "회원 마일리지 통계", "❌ 미포함", ""),
    ("7. 통계/분석", "63종 세부 통계 보고서", "❌ 미포함", "고도몰은 63종 보유"),

    # 8. 알림/커뮤니케이션
    ("8. 알림/커뮤니케이션", "카카오 알림톡 자동발송", "✅ 포함", "솔라피 직접 연동"),
    ("8. 알림/커뮤니케이션", "SMS 자동발송", "✅ 포함", "솔라피 직접 연동"),
    ("8. 알림/커뮤니케이션", "이메일 자동발송", "✅ 포함", "Resend"),
    ("8. 알림/커뮤니케이션", "080 수신거부 서비스", "❌ 미포함", "광고성 SMS 법적 의무"),
    ("8. 알림/커뮤니케이션", "알림 로그 관리", "🔶 부분포함", ""),
    ("8. 알림/커뮤니케이션", "알림 템플릿 관리", "🔶 부분포함", ""),
    ("8. 알림/커뮤니케이션", "관리자 수동 문자발송", "❌ 미포함", ""),
    ("8. 알림/커뮤니케이션", "이메일 수동 발송", "❌ 미포함", ""),
    ("8. 알림/커뮤니케이션", "이메일 충전/과금 관리", "➖ 불필요", ""),

    # 9. 게시판/콘텐츠
    ("9. 게시판/콘텐츠", "공지사항", "❌ 미포함", ""),
    ("9. 게시판/콘텐츠", "FAQ", "❌ 미포함", ""),
    ("9. 게시판/콘텐츠", "1:1 문의 (Q&A)", "❌ 미포함", ""),
    ("9. 게시판/콘텐츠", "상품 리뷰", "✅ 포함", "Phase 2"),
    ("9. 게시판/콘텐츠", "베스트 리뷰 선정", "❌ 미포함", ""),
    ("9. 게시판/콘텐츠", "리뷰 마일리지 지급", "❌ 미포함", ""),
    ("9. 게시판/콘텐츠", "자유 게시판", "❌ 미포함", ""),
    ("9. 게시판/콘텐츠", "이벤트 게시판", "❌ 미포함", ""),
    ("9. 게시판/콘텐츠", "게시판 금지어 설정", "❌ 미포함", ""),
    ("9. 게시판/콘텐츠", "팝업 관리", "❌ 미포함", ""),
    ("9. 게시판/콘텐츠", "배너 관리", "❌ 미포함", ""),

    # 10. 디자인/스킨
    ("10. 디자인/스킨", "스킨 관리", "➖ 불필요", "코드 직접 수정 방식"),
    ("10. 디자인/스킨", "페이지 에디터 (노코드)", "❌ 미포함", ""),
    ("10. 디자인/스킨", "배너 그룹 관리", "❌ 미포함", ""),
    ("10. 디자인/스킨", "팝업 관리", "❌ 미포함", ""),
    ("10. 디자인/스킨", "메인 진열 설정", "❌ 미포함", ""),
    ("10. 디자인/스킨", "네비게이션 설정", "❌ 미포함", ""),
    ("10. 디자인/스킨", "테마 관리", "❌ 미포함", ""),
    ("10. 디자인/스킨", "모바일 전용 설정", "🔶 부분포함", "Responsive로 통합"),

    # 11. SCM/공급사
    ("11. SCM/공급사", "공급사 등록/관리", "❌ 미포함", ""),
    ("11. SCM/공급사", "수수료 관리", "❌ 미포함", ""),
    ("11. SCM/공급사", "공급사 정산", "❌ 미포함", ""),
    ("11. SCM/공급사", "공급사 상품 신청", "❌ 미포함", ""),
    ("11. SCM/공급사", "공급사 세금계산서", "❌ 미포함", ""),
    ("11. SCM/공급사", "공급사 게시판", "❌ 미포함", ""),

    # 12. 외부채널 연동
    ("12. 외부채널 연동", "네이버 쇼핑 피드", "❌ 미포함", "Phase 3"),
    ("12. 외부채널 연동", "구글 쇼핑 피드", "❌ 미포함", ""),
    ("12. 외부채널 연동", "카카오 모먼트 광고", "❌ 미포함", ""),
    ("12. 외부채널 연동", "페이스북 광고 연동", "❌ 미포함", ""),
    ("12. 외부채널 연동", "네이버/다음 CPC", "❌ 미포함", ""),
    ("12. 외부채널 연동", "ShopLinker 연동", "❌ 미포함", ""),

    # 13. 시스템/보안
    ("13. 시스템/보안", "관리자 계정 다중 관리", "❌ 미포함", ""),
    ("13. 시스템/보안", "관리자 권한 세분화", "❌ 미포함", ""),
    ("13. 시스템/보안", "관리자 접근 로그", "❌ 미포함", ""),
    ("13. 시스템/보안", "SEO 설정 (메타태그)", "🔶 부분포함", "Next.js 기본"),
    ("13. 시스템/보안", "사이트맵 자동 생성", "❌ 미포함", ""),
    ("13. 시스템/보안", "외부 스크립트 관리", "❌ 미포함", ""),
    ("13. 시스템/보안", "개인정보 처리방침", "❌ 미포함", "법적 의무"),
    ("13. 시스템/보안", "이용약관 관리", "❌ 미포함", "법적 의무"),
    ("13. 시스템/보안", "파일 스토리지 관리", "🔶 부분포함", "Supabase Storage"),
    ("13. 시스템/보안", "엑셀 일괄 처리 (전체)", "❌ 미포함", ""),
    ("13. 시스템/보안", "국제화/다국어", "❌ 미포함", ""),
    ("13. 시스템/보안", "모바일 앱 설정", "❌ 미포함", ""),
    ("13. 시스템/보안", "Rate Limiting / DDoS 방어", "✅ 포함", "Upstash Redis (고도몰 없음)"),
]

# ── 데이터 입력 ──
status_fill_map = {
    "✅": FILL_YES,
    "🔶": FILL_PARTIAL,
    "❌": FILL_NO,
    "➖": FILL_NA,
}

row = 2
for cat, feature, status, note in data:
    # 카테고리 셀
    cat_cell = ws.cell(row=row, column=1, value=cat)
    cat_cell.alignment = ALIGN_CENTER
    cat_cell.font = FONT_NORMAL
    cat_cell.border = BORDER

    # 기능명
    feat_cell = ws.cell(row=row, column=2, value=feature)
    feat_cell.alignment = ALIGN_LEFT
    feat_cell.font = FONT_NORMAL
    feat_cell.border = BORDER

    # 포함여부
    status_cell = ws.cell(row=row, column=3, value=status)
    status_cell.alignment = ALIGN_CENTER
    status_cell.font = FONT_NORMAL
    status_cell.border = BORDER
    prefix = status[:2]
    for key, fill in status_fill_map.items():
        if key in prefix:
            status_cell.fill = fill
            break

    # 비고
    note_cell = ws.cell(row=row, column=4, value=note)
    note_cell.alignment = ALIGN_LEFT
    note_cell.font = FONT_NORMAL
    note_cell.border = BORDER

    # 설성목장 필요여부 (빈칸 - 사용자 입력)
    for col in [5, 6, 7]:
        c = ws.cell(row=row, column=col, value="")
        c.alignment = ALIGN_CENTER
        c.font = FONT_NORMAL
        c.border = BORDER

    row += 1

# ── 범례 시트 ──
ws2 = wb.create_sheet("범례 및 안내")
legend = [
    ("기호", "의미", "색상"),
    ("✅ 포함", "초기 기획에 포함된 기능", "연초록"),
    ("🔶 부분포함", "일부만 포함, 축소 구현", "연노랑"),
    ("❌ 미포함", "기획 미포함, 추가 개발 필요", "연빨강"),
    ("➖ 불필요", "설성목장 비즈니스에 불필요 판단", "연회색"),
]
for r, row_data in enumerate(legend, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = ALIGN_CENTER
        if r == 1:
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
        elif r == 2:
            cell.fill = FILL_YES
        elif r == 3:
            cell.fill = FILL_PARTIAL
        elif r == 4:
            cell.fill = FILL_NO
        elif r == 5:
            cell.fill = FILL_NA

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 35
ws2.column_dimensions["C"].width = 12

ws2.cell(row=7, column=1, value="안내").font = Font(bold=True)
ws2.cell(row=8, column=1, value="'설성목장 필요여부' 열에 O/X 또는 필요/불필요를 입력 후 업로드해주세요.")
ws2.merge_cells("A8:C8")

# ── 열 너비 / 행 높이 ──
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 25
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 22

ws.row_dimensions[1].height = 24
for i in range(2, row):
    ws.row_dimensions[i].height = 18

# ── 카테고리 병합 ──
from itertools import groupby
rows_list = list(ws.iter_rows(min_row=2, max_row=row-1, min_col=1, max_col=1))
cat_groups = []
start = 2
current_cat = data[0][0]
for i, (cat, *_) in enumerate(data):
    if cat != current_cat:
        cat_groups.append((current_cat, start, start + i - len([x for x in data[:i] if x[0]==current_cat]) - 1 + len([x for x in data[:i] if x[0]==current_cat])))
        current_cat = cat
        start = i + 2

# 카테고리별 색상 지정
cat_colors = [
    "1F4E79","2E75B6","2196A6","217346","375623",
    "843C0C","7030A0","C00000","833C00","1F3864",
    "4D4D4D","00B0F0","375623"
]
cat_list = list(dict.fromkeys(d[0] for d in data))
start_rows = {}
for i, (cat, *_) in enumerate(data):
    if cat not in start_rows:
        start_rows[cat] = i + 2

for cat_idx, cat_name in enumerate(cat_list):
    indices = [i+2 for i, d in enumerate(data) if d[0] == cat_name]
    s, e = min(indices), max(indices)
    color = cat_colors[cat_idx % len(cat_colors)]
    fill = PatternFill("solid", fgColor=color)
    if s < e:
        ws.merge_cells(f"A{s}:A{e}")
    cell = ws.cell(row=s, column=1)
    cell.fill = fill
    cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.alignment = ALIGN_CENTER

# ── 헤더 고정 ──
ws.freeze_panes = "A2"

# ── 자동 필터 ──
ws.auto_filter.ref = f"A1:G{row-1}"

# ── 요약 시트 ──
ws3 = wb.create_sheet("요약")
summary_headers = ["카테고리", "전체 기능수", "✅ 포함", "🔶 부분포함", "❌ 미포함", "➖ 불필요", "포함율(%)"]
for c, h in enumerate(summary_headers, 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.fill = FILL_HEADER
    cell.font = FONT_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER

from collections import Counter
summary_data = {}
for cat, feat, status, note in data:
    if cat not in summary_data:
        summary_data[cat] = Counter()
    if "✅" in status:
        summary_data[cat]["yes"] += 1
    elif "🔶" in status:
        summary_data[cat]["partial"] += 1
    elif "❌" in status:
        summary_data[cat]["no"] += 1
    elif "➖" in status:
        summary_data[cat]["na"] += 1

total_yes = total_partial = total_no = total_na = 0
for r, (cat, counts) in enumerate(summary_data.items(), 2):
    yes = counts["yes"]
    partial = counts["partial"]
    no = counts["no"]
    na = counts["na"]
    total = yes + partial + no + na
    rate = round((yes + partial * 0.5) / (total - na) * 100 if (total - na) > 0 else 0, 1)

    total_yes += yes; total_partial += partial
    total_no += no; total_na += na

    row_data = [cat, total, yes, partial, no, na, f"{rate}%"]
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
        cell.font = FONT_NORMAL
        if c == 3: cell.fill = FILL_YES
        elif c == 4: cell.fill = FILL_PARTIAL
        elif c == 5: cell.fill = FILL_NO
        elif c == 6: cell.fill = FILL_NA

# 합계 행
total_all = total_yes + total_partial + total_no + total_na
total_rate = round((total_yes + total_partial * 0.5) / (total_all - total_na) * 100, 1)
total_row = len(summary_data) + 2
totals = ["합계", total_all, total_yes, total_partial, total_no, total_na, f"{total_rate}%"]
for c, val in enumerate(totals, 1):
    cell = ws3.cell(row=total_row, column=c, value=val)
    cell.font = Font(bold=True, size=10)
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER
    cell.fill = PatternFill("solid", fgColor="D9E1F2")

for col in ["A","B","C","D","E","F","G"]:
    ws3.column_dimensions[col].width = 18
ws3.freeze_panes = "A2"

# ── 저장 ──
out = "/Users/sungjincho/Documents/Projects/설성목장몰_기능비교표.xlsx"
wb.save(out)
print(f"저장 완료: {out}")
